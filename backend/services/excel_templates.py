import io
from datetime import datetime
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# Ensure these exist in backend.utils.constants
from backend.utils.constants import (
    INSTITUTION_LIST,
    ACCOUNT_TYPES,             # kept for compatibility
    ACCOUNT_CATEGORIES,
    ACCOUNT_TYPES_BY_CATEGORY, # dict[str, list[str]]
)

class ExcelTemplateService:
    """Service for generating Excel import templates (Accounts + Positions)."""

    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        self.required_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        self.sample_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

        self.border = Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            top=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD"),
        )

        self.lookup_ranges = {"institutions": None, "categories": None, "types_flat": None}
        self.named_ranges = {
            "institutions": "InstitutionsList",
            "categories": "AccountCategoriesList",
            "types_flat": "AccountTypesList",
        }
    # =========================
    # PUBLIC: ACCOUNTS TEMPLATE
    # =========================
    def create_accounts_template(self) -> io.BytesIO:
        wb = Workbook()
        wb.remove(wb.active)

        # Create sheets in order
        self._create_instructions_sheet(wb)
        accounts_ws = self._create_accounts_sheet(wb)
        lookups_ws = self._create_validation_lists_sheet(wb)
        
        # Add validations AFTER creating lookups sheet with the workbook context
        self._add_institution_validation_named_range(accounts_ws)
        self._add_category_validation_named_range(accounts_ws)
        self._add_account_type_validation_with_indirect(accounts_ws, wb)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _add_account_type_validation_with_indirect(self, ws: Worksheet, wb: Workbook) -> None:
        """
        Add account type validation that dynamically changes based on category selection.
        Uses INDIRECT formula to reference the appropriate named range.
        """
        # Create a helper column in Lookups sheet for the mapping
        lookups_ws = wb["Lookups"]
        
        # Add mapping table starting at column M (column 13)
        lookups_ws["M1"] = "Category Display"
        lookups_ws["N1"] = "Range Name"
        lookups_ws["M1"].font = self.header_font
        lookups_ws["N1"].font = self.header_font
        
        mappings = [
            ("Brokerage", "Types_Brokerage"),
            ("Retirement", "Types_Retirement"),
            ("Cash / Banking", "Types_CashBanking"),
            ("Cryptocurrency", "Types_Cryptocurrency"),
            ("Metals", "Types_Metals")
        ]
        
        for i, (display, range_name) in enumerate(mappings, start=2):
            lookups_ws.cell(row=i, column=13, value=display).border = self.border
            lookups_ws.cell(row=i, column=14, value=range_name).border = self.border
        
        # Use VLOOKUP with INDIRECT to get the right list
        # This formula looks up the category name and gets the corresponding range name
        indirect_formula = '=INDIRECT(VLOOKUP(C2,Lookups!$M$2:$N$6,2,FALSE))'
        
        dv = DataValidation(
            type="list",
            formula1=indirect_formula,
            allow_blank=True,
            errorTitle="Select Category First",
            error="Please select an Account Category first, then choose the Account Type.",
            showErrorMessage=True,
            showInputMessage=True,
            promptTitle="Account Type",
            prompt="Select a type based on your chosen category"
        )
        # Explicitly set showDropDown to False to ensure in-cell dropdown appears
        dv.showDropDown = False  # This is counterintuitive but correct for openpyxl
        ws.add_data_validation(dv)
        dv.add("D2:D5000")

    def _create_instructions_sheet(self, wb: Workbook) -> Worksheet:
        ws = wb.create_sheet("Instructions", 0)
        ws.sheet_properties.tabColor = "1976D2"

        ws.merge_cells("A1:F1")
        ws["A1"] = "NestEgg Account Import Instructions"
        ws["A1"].font = Font(bold=True, size=16, color="2C3E50")
        ws["A1"].alignment = Alignment(horizontal="center")

        lines = [
            "",
            "Step-by-Step Guide:",
            "1. Go to the 'Accounts' tab",
            "2. Fill in your account information (yellow cells are required)",
            "3. Use the dropdown menus for Institution, Account Type, and Category",
            "4. Delete the example rows on the 'Accounts' sheet (rows 2–5) before importing",
            "5. Save this file",
            "6. Upload to NestEgg using the 'Import Accounts' feature",
            "",
            "Important Notes:",
            "• Required fields are highlighted in YELLOW",
            "• Institution, Account Category, and Account Type must be selected from the dropdowns",
            "• Account names should be unique (avoid duplicates)",
            "• Don't modify the column headers",
            "• Leave cells empty rather than typing 'N/A'",
            "",
            "Field Descriptions:",
            "• Account Name: Your chosen name for the account (e.g., 'My 401k', 'Joint Savings')",
            "• Institution: The financial institution where the account is held",
            "• Account Category: General grouping (retirement, cash, brokerage, etc.)",
            "• Account Type: Specific type within the category (e.g., 'Roth IRA', 'Checking')",
            "",
            "Need Help?:",
            "• Missing an institution? Contact support@nesteggfinapp.com",
            "• For account type questions, refer to the Category-Type Reference tab",
            "",
            "After importing accounts, you can:",
            "• Download a customized Positions template with your account names",
            "• Add securities, cash, crypto, and metals to your accounts",
        ]

        row = 3
        for text in lines:
            ws[f"A{row}"] = text
            ws[f"A{row}"].font = Font(bold=True, size=12, color="2C3E50") if text.endswith(":") else Font(size=11)
            row += 1

        ws.column_dimensions["A"].width = 100
        return ws

    def _create_accounts_sheet(self, wb: Workbook) -> Worksheet:
        ws = wb.create_sheet("Accounts", 1)
        ws.sheet_properties.tabColor = "4CAF50"

        headers = ["Account Name", "Institution", "Account Category", "Account Type", "Notes"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        samples = [
            ["My 401k", "Fidelity", "Retirement", "401(k)", ""],
            ["Joint Brokerage", "Vanguard", "Brokerage", "Joint", ""],
            ["Emergency Fund", "Ally Bank", "Cash / Banking", "High Yield Savings", ""],
            ["Bitcoin Wallet", "Coinbase", "Cryptocurrency", "Exchange Account", ""],
        ]
        for r_idx, row_vals in enumerate(samples, start=2):
            for c_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = self.border
                cell.fill = self.sample_fill

        for r in range(6, 250):
            for c in range(1, 5+1):
                cell = ws.cell(row=r, column=c, value="")
                cell.border = self.border
                if c <= 4:
                    cell.fill = self.required_fill

        widths = {"A": 30, "B": 25, "C": 22, "D": 25, "E": 40}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        note_row = 6
        ws.merge_cells(f"A{note_row}:E{note_row}")
        ws[f"A{note_row}"] = "Important: Delete the example rows (2–5) before uploading."
        ws[f"A{note_row}"].font = Font(bold=True, color="9C6500")
        ws[f"A{note_row}"].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws[f"A{note_row}"].alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"

        return ws

    def _create_validation_lists_sheet(self, wb: Workbook) -> Worksheet:
        ws = wb.create_sheet("Lookups", 2)
        # Hide this sheet from users
        ws.sheet_state = 'hidden'

        # Institutions column
        ws["A1"] = "Institutions"
        ws["A1"].font = self.header_font
        for i, inst in enumerate(INSTITUTION_LIST, start=2):
            ws.cell(row=i, column=1, value=inst).border = self.border
        ws.column_dimensions["A"].width = 32
        
        # Account Categories column
        ws["C1"] = "Account Categories"
        ws["C1"].font = self.header_font
        category_display_names = []
        row_idx = 2  # Start at row 2
        for cat in ACCOUNT_CATEGORIES:
            if cat == "real_estate":
                continue
            # Display user-friendly names
            display_name = cat.replace("_", " ").title()
            if cat == "cash":
                display_name = "Cash / Banking"
            elif cat == "cryptocurrency":
                display_name = "Cryptocurrency"
            category_display_names.append((cat, display_name))
            ws.cell(row=row_idx, column=3, value=display_name).border = self.border
            row_idx += 1  # Increment row for next category
        ws.column_dimensions["C"].width = 22
        
        # Create dynamic category-type mapping columns for INDIRECT formula use
        col_index = 5  # Start at column E
        for category, display_name in category_display_names:
            if category == "real_estate":
                continue  # Skip real estate
                
            types = ACCOUNT_TYPES_BY_CATEGORY.get(category, [])
            
            # Header for this category's types
            ws.cell(row=1, column=col_index, value=display_name).font = self.header_font
            
            # List types for this category
            for i, t in enumerate(types, start=2):
                ws.cell(row=i, column=col_index, value=t).border = self.border
            
            # Define named range for this category's types
            # The named range should match what INDIRECT will look for
            # Remove spaces and special chars from the display name for the range name
            clean_name = display_name.replace(" / ", "").replace(" ", "")
            cat_range_addr = f"Lookups!${get_column_letter(col_index)}$2:${get_column_letter(col_index)}${1 + len(types)}"
            cat_range_name = f"Types_{clean_name}"
            
            # Create the defined name
            defined_name = DefinedName(name=cat_range_name, attr_text=cat_range_addr)
            wb.defined_names.add(defined_name)
            
            ws.column_dimensions[get_column_letter(col_index)].width = 26
            col_index += 1

        return ws

    def _add_institution_validation_named_range(self, ws: Worksheet) -> None:
        # Use direct range reference like positions template does
        inst_range = f"=Lookups!$A$2:$A${1 + len(INSTITUTION_LIST)}"
        
        dv = DataValidation(
            type="list",
            formula1=inst_range,
            allow_blank=True,
            errorTitle="Invalid Institution",
            error="Please choose a valid institution from the list or type a custom name.",
            showErrorMessage=True,
            showInputMessage=True,
            promptTitle="Institution",
            prompt="Select from the list or type a custom name"
        )
        # Explicitly set showDropDown to ensure in-cell dropdown appears
        dv.showDropDown = False  # This is counterintuitive but correct for openpyxl
        ws.add_data_validation(dv)
        dv.add("B2:B5000")

    def _add_category_validation_named_range(self, ws: Worksheet) -> None:
        # Use direct range reference
        # Count actual categories excluding real_estate
        cat_count = len([c for c in ACCOUNT_CATEGORIES if c != "real_estate"])
        cat_range = f"=Lookups!$C$2:$C${1 + cat_count}"
        
        dv = DataValidation(
            type="list",
            formula1=cat_range,
            allow_blank=True,
            errorTitle="Invalid Category",
            error="Choose a valid account category from the list.",
            showErrorMessage=True,
            showInputMessage=True,
            promptTitle="Account Category",
            prompt="Select the category that best fits this account"
        )
        # Explicitly set showDropDown to ensure in-cell dropdown appears
        dv.showDropDown = False  # This is counterintuitive but correct for openpyxl
        ws.add_data_validation(dv)
        dv.add("C2:C5000")

    def _add_account_type_validation(self, ws: Worksheet) -> None:
        # Use INDIRECT formula to dynamically reference the category's type list
        # This formula will look up the named range based on the category selected in column C
        indirect_formula = '=INDIRECT("Types_" & SUBSTITUTE(SUBSTITUTE(C2," ","")," / ",""))'
        
        dv = DataValidation(
            type="list",
            formula1=indirect_formula,
            allow_blank=True,
            errorTitle="Select Category First",
            error="Please select an Account Category first, then choose the Account Type.",
            showErrorMessage=True,
            showInputMessage=True,
            promptTitle="Account Type",
            prompt="Select a type based on your chosen category"
        )
        # Explicitly set showDropDown to ensure in-cell dropdown appears
        dv.showDropDown = False  # This is counterintuitive but correct for openpyxl
        ws.add_data_validation(dv)
        dv.add("D2:D5000")

    def _create_category_type_reference(self, wb: Workbook) -> Worksheet:
        ws = wb.create_sheet("Category-Type Reference", 3)
        ws.sheet_properties.tabColor = "FFC107"
        # Hide this reference sheet from users
        ws.sheet_state = 'hidden'

        ws.merge_cells("A1:C1")
        ws["A1"] = "Valid Account Types by Category"
        ws["A1"].font = Font(bold=True, size=14, color="2C3E50")
        ws["A1"].alignment = Alignment(horizontal="center")

        row = 3
        for category, types in ACCOUNT_TYPES_BY_CATEGORY.items():
            ws.merge_cells(f"A{row}:C{row}")
            ws.cell(row=row, column=1, value=category.replace("_", " ").title())
            ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
            ws.cell(row=row, column=1).fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            for t in types:
                row += 1
                ws.cell(row=row, column=2, value=t).border = self.border
                ws.cell(row=row, column=3, value="✓").font = Font(color="27AE60", bold=True)
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
            row += 2

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 10
        ws.freeze_panes = "A2"
        return ws

    # ==========================
    # PUBLIC: POSITIONS TEMPLATE
    # (simple scaffold; Step 2)
    # ==========================

    async def create_positions_template(self, user_id: Any, database) -> io.BytesIO:
        # Get user's accounts
        accounts_query = """
            SELECT id, account_name, institution
            FROM accounts
            WHERE user_id = :user_id
            ORDER BY account_name
        """
        accounts = await database.fetch_all(query=accounts_query, values={"user_id": user_id})
        if not accounts:
            raise ValueError("No accounts found. Please create accounts before downloading positions template.")
        
        # Get securities (stocks/ETFs only)
        securities_query = """
            SELECT ticker, company_name
            FROM securities
            WHERE asset_type IN ('security', 'index')
            AND on_yfinance = true
            ORDER BY ticker
            LIMIT 5000  -- Reasonable limit for Excel dropdown
        """
        securities = await database.fetch_all(query=securities_query)
        
        # Get crypto assets
        crypto_query = """
            SELECT ticker, company_name
            FROM securities
            WHERE asset_type = 'crypto'
            ORDER BY ticker
            LIMIT 500
        """
        cryptos = await database.fetch_all(query=crypto_query)
        
        # Get metal futures
        metals_query = """
            SELECT 
                CASE 
                    WHEN ticker = 'GC=F' THEN 'Gold'
                    WHEN ticker = 'SI=F' THEN 'Silver'
                    WHEN ticker = 'PL=F' THEN 'Platinum'
                    WHEN ticker = 'PA=F' THEN 'Palladium'
                    WHEN ticker = 'HG=F' THEN 'Copper'
                END as metal_type,
                ticker,
                company_name,
                current_price
            FROM securities
            WHERE ticker IN ('GC=F', 'SI=F', 'PL=F', 'PA=F', 'HG=F')
            ORDER BY metal_type
        """
        metals = await database.fetch_all(query=metals_query)

        wb = Workbook()
        wb.remove(wb.active)
        
        # Create comprehensive instructions
        self._create_master_instructions_sheet(wb)
        
        # Create hidden lookups sheet
        lookups = self._create_lookups_sheet(wb, accounts, securities, cryptos, metals)
        
        # Create position entry sheets with clear labeling
        self._create_securities_positions_sheet(wb, accounts, securities)
        self._create_cash_positions_sheet(wb, accounts)
        self._create_crypto_positions_sheet(wb, accounts, cryptos)
        self._create_metal_positions_sheet(wb, accounts, metals)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _create_master_instructions_sheet(self, wb: Workbook) -> Worksheet:
        ws = wb.create_sheet("📋 START HERE - Instructions", 0)
        ws.sheet_properties.tabColor = "FF0000"  # Red to draw attention
        
        # Title
        ws.merge_cells("A1:H1")
        ws["A1"] = "NestEgg Position Import Template"
        ws["A1"].font = Font(bold=True, size=18, color="2C3E50")
        ws["A1"].alignment = Alignment(horizontal="center")
        
        # Overview
        ws.merge_cells("A3:H3")
        ws["A3"] = "HOW TO USE THIS TEMPLATE"
        ws["A3"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A3"].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws["A3"].alignment = Alignment(horizontal="center")
        
        # Tab descriptions
        instructions = [
            "",
            "This workbook contains 4 position type tabs. Fill in ONLY the tabs for positions you want to import:",
            "",
            "📈 Tab 1: SECURITIES (Blue) - For stocks, ETFs, mutual funds",
            "   • Required: Account, Ticker, Shares, Cost Basis, Purchase Date",
            "   • Ticker dropdown includes all tracked securities",
            "   • Company name auto-fills based on ticker selection",
            "",
            "💵 Tab 2: CASH (Purple) - For savings, checking, money market, CDs",
            "   • Required: Account, Cash Type, Amount",
            "   • Optional: Interest rate, maturity date",
            "",
            "🪙 Tab 3: CRYPTO (Orange) - For cryptocurrency holdings",
            "   • Required: Account, Symbol, Quantity, Purchase Price, Purchase Date",
            "   • Symbol dropdown includes tracked cryptocurrencies",
            "",
            "🥇 Tab 4: METALS (Gold) - For precious metal holdings",
            "   • Required: Account, Metal Type, Quantity (oz), Purchase Price, Purchase Date",
            "   • Metal type dropdown includes Gold, Silver, Platinum, etc.",
            "   • Current prices auto-update from futures markets",
            "",
            "⚠️ IMPORTANT RULES:",
            "1. Each tab works independently - you can use any combination",
            "2. Delete ALL example rows (highlighted in blue) before importing",
            "3. Yellow cells are REQUIRED fields",
            "4. Use dropdowns where provided to ensure data accuracy",
            "5. Dates must be in YYYY-MM-DD format",
            "6. Save as .xlsx format before uploading",
            "",
            "❌ COMMON MISTAKES TO AVOID:",
            "• Don't modify column headers",
            "• Don't add data to the Lookups tab (it's for reference only)",
            "• Don't type tickers manually if dropdown is available",
            "• Don't leave required fields empty",
            "• Don't include positions for accounts that don't exist",
            "",
            "📧 SUPPORT:",
            "• Can't find your ticker? Email support@nesteggfinapp.com",
            "• Import errors? Check that all required fields are filled",
            "• Account missing? Create it in NestEgg first, then re-download template"
        ]
        
        row = 5
        for line in instructions:
            ws[f"A{row}"] = line
            if line.startswith("📈") or line.startswith("💵") or line.startswith("🪙") or line.startswith("🥇"):
                ws[f"A{row}"].font = Font(bold=True, size=12, color="2C3E50")
                ws.merge_cells(f"A{row}:H{row}")
            elif line.startswith("⚠️") or line.startswith("❌") or line.startswith("📧"):
                ws[f"A{row}"].font = Font(bold=True, size=12, color="C00000")
                ws.merge_cells(f"A{row}:H{row}")
            elif line.startswith("   •"):
                ws[f"A{row}"].font = Font(size=10, color="666666")
                ws.merge_cells(f"A{row}:H{row}")
            else:
                ws[f"A{row}"].font = Font(size=11)
                ws.merge_cells(f"A{row}:H{row}")
            row += 1
        
        ws.column_dimensions["A"].width = 100
        return ws

    def _create_lookups_sheet(self, wb: Workbook, accounts: list, securities: list, cryptos: list, metals: list) -> Worksheet:
        ws = wb.create_sheet("Lookups", 1)
        ws.sheet_state = 'hidden'
        
        # Accounts column
        ws["A1"] = "Accounts"
        ws["A1"].font = self.header_font
        for i, acc in enumerate(accounts, start=2):
            ws.cell(row=i, column=1, value=acc["account_name"]).border = self.border
        
        # Create named range for accounts
        if accounts:
            acc_range = f"Lookups!$A$2:$A${1+len(accounts)}"
            defined_name = DefinedName(name="AccountsList", attr_text=acc_range)
            wb.defined_names.add(defined_name)
            
        # Securities column (ticker and company name)
        ws["C1"] = "Ticker"
        ws["D1"] = "Company"
        ws["C1"].font = self.header_font
        ws["D1"].font = self.header_font
        for i, sec in enumerate(securities, start=2):
            ws.cell(row=i, column=3, value=sec["ticker"]).border = self.border
            ws.cell(row=i, column=4, value=sec["company_name"]).border = self.border
        
        # ALWAYS create named range for securities
        sec_last = 1 + max(len(securities), 1)
        sec_range = f"Lookups!$C$2:$C${sec_last}"
        wb.defined_names.add(DefinedName(name="SecuritiesList", attr_text=sec_range))
        
        # Crypto column
        ws["F1"] = "Crypto Symbol"
        ws["G1"] = "Crypto Name"
        ws["F1"].font = self.header_font
        ws["G1"].font = self.header_font
        for i, crypto in enumerate(cryptos, start=2):
            ws.cell(row=i, column=6, value=crypto["ticker"]).border = self.border
            ws.cell(row=i, column=7, value=crypto["company_name"]).border = self.border
        
        # ALWAYS create named range for cryptos
        crypto_last = 1 + max(len(cryptos), 1)
        crypto_range = f"Lookups!$F$2:$F${crypto_last}"
        wb.defined_names.add(DefinedName(name="CryptoList", attr_text=crypto_range))
        
        # Metals column
        ws["I1"] = "Metal Type"
        ws["J1"] = "Metal Ticker"
        ws["I1"].font = self.header_font
        ws["J1"].font = self.header_font
        if metals:
            for i, metal in enumerate(metals, start=2):
                ws.cell(row=i, column=9, value=metal["metal_type"]).border = self.border
                ws.cell(row=i, column=10, value=metal["ticker"]).border = self.border
        
        # ALWAYS create named range for metals
        metal_last = 1 + max(len(metals), 1)
        metal_range = f"Lookups!$I$2:$I${metal_last}"
        wb.defined_names.add(DefinedName(name="MetalsList", attr_text=metal_range))
        
        # Cash types
        ws["L1"] = "Cash Types"
        ws["L1"].font = self.header_font
        cash_types = ["Savings", "Checking", "Money Market", "CD"]
        for i, cash_type in enumerate(cash_types, start=2):
            ws.cell(row=i, column=12, value=cash_type).border = self.border
        
        # Create named range for cash types
        cash_range = f"Lookups!$L$2:$L${1+len(cash_types)}"
        defined_name = DefinedName(name="CashTypesList", attr_text=cash_range)
        wb.defined_names.add(defined_name)
        
        return ws

    def _create_securities_positions_sheet(self, wb: Workbook, accounts: list, securities: list) -> Worksheet:
        ws = wb.create_sheet("📈 SECURITIES", 2)
        ws.sheet_properties.tabColor = "0066CC"
        
        # Add a header banner
        ws.merge_cells("A1:G1")
        ws["A1"] = "SECURITIES POSITIONS - Stocks, ETFs, Mutual Funds"
        ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center")
        
        # Column headers (row 2)
        headers = ["Account", "Ticker", "Company Name", "Shares", "Cost Basis", "Purchase Date"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Set column widths
        widths = {"A": 25, "B": 15, "C": 35, "D": 15, "E": 15, "F": 15, "G": 30}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        
        # Add account dropdown validation
        if accounts:
            dv_acc = DataValidation(
                type="list",
                formula1="=AccountsList",
                allow_blank=False
            )
            dv_acc.showDropDown = False  # Force dropdown arrow to appear
            ws.add_data_validation(dv_acc)
            dv_acc.add("A3:A1000")

        # Add ticker dropdown validation  
        if securities:
            dv_ticker = DataValidation(
                type="list",
                formula1="=SecuritiesList",
                allow_blank=False
            )
            dv_ticker.showDropDown = False  # Force dropdown arrow to appear
            ws.add_data_validation(dv_ticker)
            dv_ticker.add("B3:B1000")
            # Auto-fill Company Name for ALL editable rows with a VLOOKUP on the ticker
            for r in range(3, 1001):
                ws.cell(row=r, column=3, value=f'=IFERROR(VLOOKUP(B{r},Lookups!$C:$D,2,FALSE),"")').border = self.border

        
        # Add numeric validation for shares
        dv_shares = DataValidation(
            type="decimal",
            operator="greaterThan",
            formula1=0,
            allow_blank=False
        )
        ws.add_data_validation(dv_shares)
        dv_shares.add("D3:D1000")
        
        # Add numeric validation for cost basis
        dv_cost = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1=0,
            allow_blank=False
        )
        ws.add_data_validation(dv_cost)
        dv_cost.add("E3:E1000")
        
        # Add date validation
        dv_date = DataValidation(
            type="date",
            operator="between",
            formula1="DATE(1900,1,1)",
            formula2="TODAY()",
            allow_blank=False
        )
        ws.add_data_validation(dv_date)
        dv_date.add("F3:F1000")
        
        # Add example rows
        examples = [
            ["Select Account", "AAPL", "=IFERROR(VLOOKUP(B3,Lookups!$C:$D,2,FALSE),\"\")", "100", "150.00", "2024-01-15"],
            ["Select Account", "VOO", "=IFERROR(VLOOKUP(B4,Lookups!$C:$D,2,FALSE),\"\")", "50", "425.00", "2024-02-01"]
        ]
        
        for r_idx, row_data in enumerate(examples, start=3):
            for c_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = self.border
                cell.fill = self.sample_fill
        
        # Required field highlighting for empty rows
        for row in range(5, 50):
            for col in [1, 2, 4, 5, 6]:  # Required columns
                ws.cell(row=row, column=col).fill = self.required_fill
        
        # Add instruction row — put directly under examples on row 5, and remove '*' language
        ws.merge_cells("A5:F5")
        ws["A5"] = "⚠️ Delete example rows (3–4) before importing."
        ws["A5"].font = Font(bold=True, italic=True, color="FF0000", size=10)
        ws["A5"].alignment = Alignment(horizontal="center")
        
        return ws

    def _create_cash_positions_sheet(self, wb: Workbook, accounts: list) -> Worksheet:
        ws = wb.create_sheet("💵 CASH", 3)
        ws.sheet_properties.tabColor = "663399"
        
        # Header banner
        ws.merge_cells("A1:F1")
        ws["A1"] = "CASH POSITIONS - Savings, Checking, Money Market, CDs"
        ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="663399", end_color="663399", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ["Account", "Cash Type", "Amount", "Interest Rate (%)", "Maturity Date", "Notes"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Column widths
        widths = {"A": 25, "B": 20, "C": 15, "D": 15, "E": 15, "F": 30}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        
        # Account validation
        # Account validation
        if accounts:
            dv_acc = DataValidation(
                type="list", 
                formula1="=AccountsList",  # Use named range
                allow_blank=False
            )
            dv_acc.showDropDown = False
            ws.add_data_validation(dv_acc)
            dv_acc.add("A3:A1000")

        # Cash type validation
        dv_type = DataValidation(
            type="list", 
            formula1="=CashTypesList",  # Use named range
            allow_blank=False
        )
        dv_type.showDropDown = False  # ensure in-cell dropdown shows
        ws.add_data_validation(dv_type)
        dv_type.add("B3:B1000")

        # Amount validation
        dv_amount = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1=0,
            allow_blank=False
        )
        ws.add_data_validation(dv_amount)
        dv_amount.add("C3:C1000")

        
        # Example rows
        examples = [
            ["Select Account", "Savings", "10000", "4.5", "", "Emergency fund"],
            ["Select Account", "CD", "25000", "5.0", "2025-06-30", "6-month CD"]
        ]
        
        for r_idx, row_data in enumerate(examples, start=3):
            for c_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = self.border
                cell.fill = self.sample_fill
        
        # Required fields highlighting
        for row in range(5, 50):
            for col in [1, 2, 3]:  # Required columns
                ws.cell(row=row, column=col).fill = self.required_fill

        # Add instruction row — put directly under examples on row 5, and remove '*' language
        ws.merge_cells("A5:F5")
        ws["A5"] = "⚠️ Delete example rows (3–4) before importing."
        ws["A5"].font = Font(bold=True, italic=True, color="FF0000", size=10)
        ws["A5"].alignment = Alignment(horizontal="center")
        
        return ws

    def _create_crypto_positions_sheet(self, wb: Workbook, accounts: list, cryptos: list) -> Worksheet:
        ws = wb.create_sheet("🪙 CRYPTO", 4)
        ws.sheet_properties.tabColor = "FF9900"
        
        # Header banner
        ws.merge_cells("A1:G1")
        ws["A1"] = "CRYPTOCURRENCY POSITIONS"
        ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ["Account", "Symbol", "Name", "Quantity", "Purchase Price", "Purchase Date"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Column widths
        widths = {"A": 25, "B": 15, "C": 35, "D": 20, "E": 20, "F": 15, "G": 20}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        
        # Account validation (use named range for consistency)
        if accounts:
            dv_acc = DataValidation(type="list", formula1="=AccountsList", allow_blank=False)
            dv_acc.showDropDown = False
            ws.add_data_validation(dv_acc)
            dv_acc.add("A3:A1000")

        # Crypto symbol validation (always add; named range exists even if empty)
        dv_crypto = DataValidation(
            type="list", 
            formula1="=CryptoList",
            allow_blank=False
        )
        dv_crypto.showDropDown = False
        ws.add_data_validation(dv_crypto)
        dv_crypto.add("B3:B1000")
        for r in range(3, 1001):
            ws.cell(row=r, column=3, value=f'=IFERROR(VLOOKUP(B{r},Lookups!$F:$G,2,FALSE),"")').border = self.border




        # Quantity validation
        dv_qty = DataValidation(
            type="decimal",
            operator="greaterThan",
            formula1=0,
            allow_blank=False
        )
        ws.add_data_validation(dv_qty)
        dv_qty.add("D3:D1000")
        
        # Price validation
        dv_price = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1=0,
            allow_blank=False
        )
        ws.add_data_validation(dv_price)
        dv_price.add("E3:E1000")
        
        dv_date = DataValidation(
            type="date",
            operator="between",
            formula1="DATE(1900,1,1)",
            formula2="TODAY()",
            allow_blank=False
        )
        ws.add_data_validation(dv_date)
        dv_date.add("F3:F1000")
        
        # Example rows
        examples = [
            ["Select Account", "BTC", "=IFERROR(VLOOKUP(B3,Lookups!$F:$G,2,FALSE),\"\")", "0.5", "45000", "2024-01-15"],
            ["Select Account", "ETH", "=IFERROR(VLOOKUP(B4,Lookups!$F:$G,2,FALSE),\"\")", "2.0", "2500", "2024-02-01"]
        ]
        
        for r_idx, row_data in enumerate(examples, start=3):
            for c_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = self.border
                cell.fill = self.sample_fill
        
        # Required fields highlighting
        for row in range(5, 50):
            for col in [1, 2, 4, 5, 6]:  # Required columns
                ws.cell(row=row, column=col).fill = self.required_fill

        # Add instruction row — put directly under examples on row 5, and remove '*' language
        ws.merge_cells("A5:F5")
        ws["A5"] = "⚠️ Delete example rows (3–4) before importing."
        ws["A5"].font = Font(bold=True, italic=True, color="FF0000", size=10)
        ws["A5"].alignment = Alignment(horizontal="center")
        
        return ws

    def _create_metal_positions_sheet(self, wb: Workbook, accounts: list, metals: list) -> Worksheet:
        ws = wb.create_sheet("🥇 METALS", 5)
        ws.sheet_properties.tabColor = "FFD700"
        
        # Header banner
        ws.merge_cells("A1:G1")
        ws["A1"] = "PRECIOUS METALS POSITIONS"
        ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ["Account", "Metal Type", "Metal Code", "Quantity (oz)", "Purchase Price/oz", "Purchase Date"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Column widths
        widths = {"A": 25, "B": 15, "C": 18, "D": 20, "E": 15, "F": 25, "G": 30}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        
        # Account validation (named range)
        if accounts:
            dv_acc = DataValidation(
                type="list", 
                formula1="=AccountsList",
                allow_blank=False
            )
            dv_acc.showDropDown = False
            ws.add_data_validation(dv_acc)
            dv_acc.add("A3:A1000")

        # Metal type validation (always add; named range exists even if empty)
        dv_metal = DataValidation(
            type="list", 
            formula1="=MetalsList",
            allow_blank=False
        )
        dv_metal.showDropDown = False 
        ws.add_data_validation(dv_metal)
        dv_metal.add("B3:B1000")
        for r in range(3, 1001):
            ws.cell(row=r, column=3, value=f'=IFERROR(VLOOKUP(B{r},Lookups!$I:$J,2,FALSE),"")').border = self.border
        
        # Quantity validation
        dv_qty = DataValidation(type="decimal", operator="greaterThan", formula1=0, allow_blank=False)
        ws.add_data_validation(dv_qty)
        dv_qty.add("D3:D1000")

        
        # Price validation       
        dv_price = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=0, allow_blank=False)
        ws.add_data_validation(dv_price)
        dv_price.add("E3:E1000")


        # Date validation      
        dv_date = DataValidation(type="date", operator="between", formula1="DATE(1900,1,1)", formula2="TODAY()", allow_blank=False)
        ws.add_data_validation(dv_date)
        dv_date.add("F3:F1000")

        # Example rows
        examples = [
            ["Select Account", "Gold",  '=IFERROR(VLOOKUP(B3,Lookups!$I:$J,2,FALSE),"")',  "10",  "1800", "2024-01-15"],
            ["Select Account", "Silver",'=IFERROR(VLOOKUP(B4,Lookups!$I:$J,2,FALSE),"")', "100",  "25",   "2024-02-01"]
        ]

        
        for r_idx, row_data in enumerate(examples, start=3):
            for c_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = self.border
                cell.fill = self.sample_fill
        
        # Required fields highlighting
        for row in range(5, 50):
            for col in [1, 2, 3, 4, 5]:  # Required columns
                ws.cell(row=row, column=col).fill = self.required_fill

        # Add instruction row — put directly under examples on row 5, and remove '*' language
        ws.merge_cells("A5:F5")
        ws["A5"] = "⚠️ Delete example rows (3–4) before importing."
        ws["A5"].font = Font(bold=True, italic=True, color="FF0000", size=10)
        ws["A5"].alignment = Alignment(horizontal="center")
        
        return ws    